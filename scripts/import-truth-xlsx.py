#!/usr/bin/env python3
"""Import the real 4-year customer book (The_PawSpace_TRUTH.xlsx → 'Customer 360 All Years') into
canonical D1 SQL. Two modes:

  --mode masked (default)  For STAGING/UAT. Real segments, revenue, dormancy and service mix are kept
                           EXACTLY (reports show true business numbers); only contact fields are masked:
                           phones become non-dialable unique placeholders (0000xxxxx) and names shorten
                           to "First L.". Marketing consent is enabled only for masked rows (safe: the
                           numbers cannot be dialled), so outbound/targeting modules light up for UAT.
  --mode live              For PRODUCTION go-live after UAT. Real names and phones. Marketing consent
                           stays 0 — under DPDP, consent must be captured fresh in the new system.

History fidelity: the workbook is customer-level (orders, gross, per-service counts, first/last dates),
so per-customer bookings are synthesized to match it exactly — order COUNT exact, gross REVENUE exact
(amount = gross/orders), SERVICE MIX from the per-service order columns, dates spread first→last. Every
insert is INSERT OR IGNORE (re-runnable). The generated SQL contains customer data — do NOT commit it;
load it and delete it.

Usage:
  python3 scripts/import-truth-xlsx.py /path/to/The_PawSpace_TRUTH.xlsx --out staging-real-seed.sql
  npx wrangler d1 execute pawspace-staging --remote --file=staging-real-seed.sql
"""
import argparse, re, sys
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl first")

def esc(s):
    return str(s or "").replace("'", "''")

def digits_phone(raw):
    m = re.search(r"\d{10,}", re.sub(r"[^0-9]", "", str(raw or "")))
    return m.group(0)[-10:] if m else ""

def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.replace(tzinfo=timezone.utc)
    s = str(v).strip()[:10]
    try: return datetime.strptime(s, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError: return None

def num(v, d=0.0):
    try: return float(v)
    except (TypeError, ValueError): return d

SVC = [("Grooming Orders", "grooming"), ("Grooming Subscription Orders", "grooming"),
       ("Training Orders", "dog_training"), ("Boarding Orders", "boarding"), ("Pet Sitting Orders", "pet_sitting")]
PRIMARY_MAP = {"grooming": "grooming", "training": "dog_training", "boarding": "boarding",
               "pet sitting": "pet_sitting", "dog walking": "dog_walking", "pet taxi": "pet_taxi"}

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx"); ap.add_argument("--mode", choices=["masked", "live"], default="masked")
    ap.add_argument("--out", default="staging-real-seed.sql"); ap.add_argument("--city", default="blr")
    a = ap.parse_args()
    masked = a.mode == "masked"
    wb = openpyxl.load_workbook(a.xlsx, read_only=True, data_only=True)
    ws = wb["Customer 360 All Years"]
    hdr = [str(h or "") for h in next(ws.iter_rows(max_row=1, values_only=True))]
    col = {h: i for i, h in enumerate(hdr)}
    pitches = {}
    if "Outbound Top 5000 Task" in wb.sheetnames:
        ows = wb["Outbound Top 5000 Task"]
        oh = [str(h or "") for h in next(ows.iter_rows(max_row=1, values_only=True))]
        oc = {h: i for i, h in enumerate(oh)}
        for r in ows.iter_rows(min_row=2, values_only=True):
            if r and r[oc.get("Customer Key", 0)] is not None:
                pitches[str(r[oc["Customer Key"]])] = str(r[oc.get("Suggested Pitch", 0)] or "")

    out = ["-- PawSpace real-customer import (%s mode). Generated file — contains customer data, do NOT commit." % a.mode,
           "CREATE TABLE IF NOT EXISTS canonical_customers (id TEXT PRIMARY KEY,city_id TEXT NOT NULL,name TEXT NOT NULL,primary_phone TEXT NOT NULL,secondary_phone TEXT,email TEXT,source TEXT NOT NULL DEFAULT 'customer_app',consent_json TEXT NOT NULL DEFAULT '{}',created_at INTEGER NOT NULL,updated_at INTEGER NOT NULL);",
           "CREATE TABLE IF NOT EXISTS crm_contacts (id TEXT PRIMARY KEY, name TEXT NOT NULL, primary_phone TEXT NOT NULL, secondary_phone TEXT, email TEXT, area TEXT, pet_names TEXT, pet_summary TEXT, stage TEXT NOT NULL DEFAULT 'New lead', owner TEXT DEFAULT 'Unassigned', source TEXT DEFAULT 'Website', lifetime_value REAL DEFAULT 0, next_action TEXT, opportunity TEXT, created_at INTEGER NOT NULL, updated_at INTEGER NOT NULL);",
           "CREATE TABLE IF NOT EXISTS customer_contact_preferences (customer_id TEXT PRIMARY KEY,marketing_consent INTEGER NOT NULL DEFAULT 0,service_consent INTEGER NOT NULL DEFAULT 1,whatsapp_consent INTEGER NOT NULL DEFAULT 0,sms_consent INTEGER NOT NULL DEFAULT 0,email_consent INTEGER NOT NULL DEFAULT 0,opt_out INTEGER NOT NULL DEFAULT 0,source TEXT NOT NULL DEFAULT 'customer',updated_by TEXT NOT NULL,updated_at INTEGER NOT NULL);",
           "CREATE TABLE IF NOT EXISTS canonical_bookings (id TEXT PRIMARY KEY,idempotency_key TEXT NOT NULL UNIQUE,customer_id TEXT NOT NULL,pet_ids_json TEXT NOT NULL,source_pet_ids_json TEXT NOT NULL,city_id TEXT NOT NULL,zone_id TEXT NOT NULL,service_code TEXT NOT NULL,package_code TEXT NOT NULL,package_name TEXT NOT NULL,schedule_group_id TEXT NOT NULL UNIQUE,provider_id TEXT NOT NULL,scheduled_start TEXT NOT NULL,scheduled_end TEXT NOT NULL,status TEXT NOT NULL DEFAULT 'confirmed',channel TEXT NOT NULL DEFAULT 'customer_app',total_amount REAL NOT NULL,currency TEXT NOT NULL DEFAULT 'INR',pricing_json TEXT NOT NULL DEFAULT '{}',created_by TEXT NOT NULL,created_at INTEGER NOT NULL,updated_at INTEGER NOT NULL);",
           "CREATE TABLE IF NOT EXISTS booking_payments (id TEXT PRIMARY KEY,booking_id TEXT NOT NULL UNIQUE,customer_id TEXT NOT NULL,amount REAL NOT NULL,amount_due_now REAL NOT NULL,currency TEXT NOT NULL DEFAULT 'INR',method TEXT NOT NULL,mode TEXT NOT NULL,status TEXT NOT NULL,gateway TEXT NOT NULL DEFAULT 'uat_sandbox',idempotency_key TEXT NOT NULL UNIQUE,detail_json TEXT NOT NULL DEFAULT '{}',created_at INTEGER NOT NULL,updated_at INTEGER NOT NULL);"]

    n = bookings = 0
    total_gross = 0.0
    for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not r or r[col["Customer Name"]] is None: continue
        n += 1
        key = str(r[col.get("Customer Key", 0)] or f"row{idx}")
        cid = "REAL-%05d" % idx
        real_name = str(r[col["Customer Name"]] or "Customer").strip()
        real_phone = digits_phone(r[col["Phone(s)"]])
        if masked:
            parts = real_name.split()
            name = parts[0] + (" " + parts[1][0] + "." if len(parts) > 1 else "")
            phone = "0000%05d" % idx  # non-dialable, unique, deterministic
        else:
            name, phone = real_name, (real_phone or "0000%05d" % idx)
        seg, pri = esc(r[col["Segment"]]), esc(r[col["Outbound Priority"]])
        nba = esc(r[col["Next Best Action"]]); pitch = esc(pitches.get(key, ""))
        first = parse_date(r[col["First Service Date"]]); last = parse_date(r[col["Last Service Date"]]) or first
        first = first or last
        orders = int(num(r[col["Orders"]])); gross = num(r[col["Gross Sales"]]); total_gross += gross
        created = int((first.timestamp() if first else 1648771200) * 1000)
        updated = int((last.timestamp() if last else 1648771200) * 1000)
        consent = 1 if (masked and str(pri).strip()) else 0
        out.append("INSERT OR IGNORE INTO canonical_customers (id,city_id,name,primary_phone,source,consent_json,created_at,updated_at) VALUES ('%s','%s','%s','%s','truth_import','{\"segment\":\"%s\"}',%d,%d);" % (cid, a.city, esc(name), phone, seg, created, updated))
        out.append("INSERT OR IGNORE INTO crm_contacts (id,name,primary_phone,stage,owner,source,lifetime_value,next_action,opportunity,created_at,updated_at) VALUES ('%s','%s','%s','Active customer','Unassigned','truth_import',%0.2f,'%s','%s',%d,%d);" % (cid, esc(name), phone, gross, nba, pitch or pri, created, updated))
        out.append("INSERT OR IGNORE INTO customer_contact_preferences (customer_id,marketing_consent,updated_by,updated_at) VALUES ('%s',%d,'truth_import',%d);" % (cid, consent, updated))
        if orders < 1 or not first: continue
        # synthesize bookings: exact count, exact gross, service mix from per-service columns
        mix = []
        for h, code in SVC:
            c = int(num(r[col.get(h, -1)] if h in col else 0))
            mix += [code] * max(0, c)
        primary = PRIMARY_MAP.get(str(r[col.get("Primary Service", 0)] or "").strip().lower(), "grooming")
        mix += [primary] * max(0, orders - len(mix))
        mix = mix[:orders]
        span = max(1, (updated - created))
        amt = round(gross / orders, 2)
        rem = round(gross - amt * orders, 2)  # rounding remainder → add to last booking
        for b, svc in enumerate(mix):
            at = created + (span * b) // max(1, orders - 1) if orders > 1 else created
            iso = datetime.fromtimestamp(at / 1000, tz=timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
            amount = amt + (rem if b == orders - 1 else 0)
            bid = "RBK-%05d-%03d" % (idx, b)
            out.append("INSERT OR IGNORE INTO canonical_bookings (id,idempotency_key,customer_id,pet_ids_json,source_pet_ids_json,city_id,zone_id,service_code,package_code,package_name,schedule_group_id,provider_id,scheduled_start,scheduled_end,status,channel,total_amount,currency,pricing_json,created_by,created_at,updated_at) VALUES ('%s','i-%s','%s','[]','[]','%s','%s-east','%s','%s-std','%s history','SG-%s','PROV-HIST','%s','%s','completed','import',%0.2f,'INR','{}','truth_import',%d,%d);" % (bid, bid, cid, a.city, a.city, svc, svc, svc, bid, iso, iso, amount, at, at))
            out.append("INSERT OR IGNORE INTO booking_payments (id,booking_id,customer_id,amount,amount_due_now,method,mode,status,idempotency_key,created_at,updated_at) VALUES ('P-%s','%s','%s',%0.2f,%0.2f,'card','prepaid','captured','pi-%s',%d,%d);" % (bid, bid, cid, amount, amount, bid, at, at))
            bookings += 1
    out.append("-- customers=%d bookings=%d gross=%0.2f mode=%s" % (n, bookings, total_gross, a.mode))
    with open(a.out, "w") as f: f.write("\n".join(out) + "\n")
    print("Wrote %s — %d customers, %d bookings, gross ₹%0.0f (%s mode)" % (a.out, n, bookings, total_gross, a.mode))

if __name__ == "__main__":
    main()
